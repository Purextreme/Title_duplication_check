import pandas as pd
from openpyxl import load_workbook
import os
from datetime import datetime
from tqdm import tqdm
from progress_bar import ProgressTracker

class ExcelProcessor:
    def __init__(self, word_processor):
        """
        初始化Excel处理器
        :param word_processor: WordProcessor实例
        """
        self.word_processor = word_processor
        self.log_entries = []
        
    def get_column_index(self, sheet, column_name):
        """
        获取列索引（不区分大小写）
        """
        # 打印调试信息
        print(f"\n正在查找列: {column_name}")
        print("当前工作表的列名:")
        for idx, cell in enumerate(sheet[1]):
            if cell.value:
                print(f"列 {idx+1}: {cell.value} (类型: {type(cell.value)})")
        
        # 将目标列名转换为小写
        column_name_lower = str(column_name).lower()
        
        # 查找列
        for idx, cell in enumerate(sheet[1]):
            if cell.value:
                cell_value_lower = str(cell.value).lower()
                if cell_value_lower == column_name_lower:
                    print(f"找到列 '{column_name}' 在位置 {idx+1}")
                    return idx + 1
                    
        print(f"未找到列 '{column_name}'")
        return None
        
    def write_log(self, log_file, message):
        """
        写入日志
        """
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        log_entry = f'[{timestamp}] {message}'
        self.log_entries.append(log_entry)
        with open(log_file, 'a', encoding='utf-8') as f:
            f.write(log_entry + '\n')
            
    def format_duplicate_info(self, duplicates):
        """
        格式化重复词信息
        """
        if not duplicates:
            return ''
            
        result = []
        for word, info in duplicates.items():
            original_forms = info['original_forms']
            if len(original_forms) > 1:
                forms_str = ', '.join(original_forms)
                result.append(f"{forms_str}: {info['count']}次")
            else:
                result.append(f"{word}: {info['count']}次")
                
        return '; '.join(result)
        
    def get_cell_value(self, cell):
        """
        获取单元格的值
        """
        return str(cell.value) if cell.value is not None else None

    def process_worksheet(self, sheet, df, config):
        """
        处理工作表
        :param sheet: 工作表对象
        :param df: DataFrame对象
        :param config: 配置信息
        """
        # 获取配置信息
        title_col = config['column_settings']['title_column']
        duplicate_col = config['column_settings']['duplicate_column']
        language_col = config['column_settings'].get('language_column')
        log_file = config.get('log_file')
        
        # 测试模式设置
        test_mode = config['check_settings'].get('test_mode', False)
        test_rows = config['check_settings'].get('test_rows', 500)
        
        # 获取列索引
        title_idx = self.get_column_index(sheet, title_col)
        if not title_idx:
            raise ValueError(f'未找到标题列 "{title_col}"')
            
        duplicate_idx = self.get_column_index(sheet, duplicate_col)
        if not duplicate_idx:
            raise ValueError(f'未找到重复词列 "{duplicate_col}"')
            
        language_idx = None
        if language_col:
            language_idx = self.get_column_index(sheet, language_col)
            if not language_idx:
                if log_file:
                    self.write_log(log_file, f'警告：未找到语言列 "{language_col}"')
        
        # 设置样式
        from openpyxl.styles import PatternFill, Font
        red_fill = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
        red_font = Font(color="FF0000")
        
        # 计算要处理的行数
        max_row = min(sheet.max_row + 1, test_rows + 1) if test_mode else sheet.max_row + 1
        valid_rows = sum(1 for row in range(2, max_row) 
                        if sheet.cell(row, title_idx).value is not None)
                        
        if test_mode:
            print(f"测试模式: 将处理 {valid_rows} 行数据")
        
        has_duplicates_found = False  # 标记是否找到重复
        
        with tqdm(total=valid_rows, desc=f"{sheet.title} 处理进度") as pbar:
            # 处理每一行
            for row in range(2, max_row):
                title = sheet.cell(row, title_idx).value
                if not title:
                    continue
                    
                # 获取语言，默认为英语
                language = 'english'
                if language_idx:
                    lang_cell = sheet.cell(row, language_idx)
                    lang_value = self.get_cell_value(lang_cell)
                    if lang_value:
                        lang_value = lang_value.strip()
                        # 根据中文判断语言
                        if lang_value == '法语':
                            language = 'french'
                        elif lang_value == '西班牙语':
                            language = 'spanish'
                        elif lang_value == '意大利语':
                            language = 'italian'
                        elif lang_value == '德语':
                            language = 'german'
                    
                if test_mode:
                    print(f"行 {row} 的语言: {language} (原始值: {lang_value if lang_value else 'None'})")

                # 处理标题
                if language == 'english':
                    has_duplicates, duplicates = self.word_processor.process_english_title(title)
                else:
                    has_duplicates, duplicates = self.word_processor.process_non_english_title(title, language)
                    
                # 更新单元格
                if has_duplicates:
                    has_duplicates_found = True
                    duplicate_info = self.format_duplicate_info(duplicates)
                    # 更新重复词列
                    duplicate_cell = sheet.cell(row, duplicate_idx)
                    duplicate_cell.value = duplicate_info
                    # 标记重复词单元格为红色
                    duplicate_cell.fill = red_fill
                    duplicate_cell.font = red_font
                    # 标记原标题单元格为红色背景
                    title_cell = sheet.cell(row, title_idx)
                    title_cell.fill = red_fill
                    if log_file:
                        self.write_log(log_file, f'行 {row}: 发现重复词 - {duplicate_info}')
                
                # 写入调试日志
                if log_file:
                    self.word_processor.write_debug_log(log_file)
                
                # 更新进度条
                pbar.update(1)
                
        self.has_duplicates = has_duplicates_found 